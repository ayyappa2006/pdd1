from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from pathlib import Path

appium_cases = [
    "Verify app launches successfully.",
    "Verify splash screen appears on startup.",
    "Verify login screen loads.",
    "Verify valid user can log in.",
    "Verify invalid credentials show an error.",
    "Verify signup screen is accessible from login.",
    "Verify signup form fields display correctly.",
    "Verify new user can navigate to signup.",
    "Verify forgot password link navigates to reset screen.",
    "Verify valid registered user can reset password.",
    "Verify dashboard displays report count.",
    "Verify dashboard cleanliness card is visible.",
    "Verify upload photo button opens upload screen.",
    "Verify gallery upload option is visible.",
    "Verify camera upload option is visible.",
    "Verify contact number field accepts input.",
    "Verify category spinner contains valid categories.",
    "Verify description field accepts text.",
    "Verify send report button is displayed.",
    "Verify submitting a report triggers confirmation.",
    "Verify report history list loads.",
    "Verify report details page opens for an entry.",
    "Verify issue status text is visible in details.",
    "Verify back navigation returns to the previous screen.",
    "Verify map view is displayed on report screen.",
    "Verify address display updates when location is found.",
    "Verify app grants required permissions at runtime.",
    "Verify auto-grant permissions works on launch.",
    "Verify empty description validation is handled.",
    "Verify invalid phone format is handled.",
    "Verify report count increments after submission.",
    "Verify user dashboard bottom nav works: History.",
    "Verify user dashboard bottom nav works: Pending.",
    "Verify user dashboard bottom nav works: Profile.",
    "Verify upload photo screen can be reopened.",
    "Verify report submission does not crash the app.",
    "Verify user profile section loads.",
    "Verify user help option is accessible.",
    "Verify user privacy page is accessible.",
    "Verify organization login screen loads.",
    "Verify org login valid credentials open dashboard.",
    "Verify org login invalid credentials show error.",
    "Verify org dashboard loads summary cards.",
    "Verify org issues screen displays issue list.",
    "Verify org issue detail view loads.",
    "Verify org can mark an issue completed.",
    "Verify org can delete an issue if supported.",
    "Verify org settings screen is accessible.",
    "Verify org help page loads.",
    "Verify org completed issues screen loads."
]

selenium_cases = [
    "Verify web login page loads.",
    "Verify login page title is correct.",
    "Verify email field is present on login.",
    "Verify password field is present on login.",
    "Verify forgot password link is present.",
    "Verify signup link is present on login page.",
    "Verify clicking signup opens signup page.",
    "Verify signup page title is correct.",
    "Verify signup name field is present.",
    "Verify signup email field is present.",
    "Verify signup password field is present.",
    "Verify signup page can navigate back to login.",
    "Verify forgot password page title is correct.",
    "Verify forgot password email field is present.",
    "Verify forgot password new password field is present.",
    "Verify forgot password back link works.",
    "Verify options page loads successfully.",
    "Verify continue as user button is visible.",
    "Verify continue as organization button is visible.",
    "Verify options user button navigates to login.",
    "Verify options org button navigates to org login.",
    "Verify user dashboard page loads by URL.",
    "Verify user history page loads by URL.",
    "Verify user AI page loads by URL.",
    "Verify user profile page loads by URL.",
    "Verify organizers list page loads by URL.",
    "Verify organizers list filter field is present.",
    "Verify org dashboard page loads by URL.",
    "Verify org issues page loads by URL.",
    "Verify org completed page loads by URL.",
    "Verify org settings page loads by URL.",
    "Verify messages list page loads by URL.",
    "Verify chat page loads by URL.",
    "Verify user navigation links exist on chat page.",
    "Verify org navigation links exist on messages page.",
    "Verify report photo page loads by URL.",
    "Verify upload photo page loads by URL.",
    "Verify homepage splash redirects to options.",
    "Verify login form submit button is present.",
    "Verify signup form submit button is present.",
    "Verify forgot password form submit button is present.",
    "Verify organizers list page displays organizers container.",
    "Verify user dashboard page displays report cards.",
    "Verify user AI page displays pending items.",
    "Verify user history page displays history items.",
    "Verify org dashboard page displays statistics.",
    "Verify org completed page displays closed issues.",
    "Verify org settings page displays profile fields.",
    "Verify web pages use correct window.location navigation.",
    "Verify static navigation links have working hrefs."
]

report_dir = Path("reports")
report_dir.mkdir(exist_ok=True)
report_path = report_dir / "CivicBin_End2End_Test_Report.xlsx"

workbook = Workbook()

# Appium sheet
appium_sheet = workbook.active
appium_sheet.title = "Appium Tests"
appium_sheet.append(["Test Case ID", "Description", "Type", "Status", "Notes"])
for idx, description in enumerate(appium_cases, start=1):
    appium_sheet.append([f"A-{idx:03}", description, "Appium", "Not Executed", ""])

# Selenium sheet
selenium_sheet = workbook.create_sheet(title="Selenium Tests")
selenium_sheet.append(["Test Case ID", "Description", "Type", "Status", "Notes"])
for idx, description in enumerate(selenium_cases, start=1):
    selenium_sheet.append([f"S-{idx:03}", description, "Selenium", "Not Executed", ""])

# Summary sheet
summary_sheet = workbook.create_sheet(title="Summary")
summary_sheet.append(["Category", "Total Tests", "Pass", "Fail", "Not Executed", "Pass Rate"])
summary_sheet.append(["Appium", len(appium_cases), 0, 0, len(appium_cases), "0%"])
summary_sheet.append(["Selenium", len(selenium_cases), 0, 0, len(selenium_cases), "0%"])
summary_sheet.append(["Total", len(appium_cases) + len(selenium_cases), 0, 0, len(appium_cases) + len(selenium_cases), "0%"])

for sheet in [appium_sheet, selenium_sheet, summary_sheet]:
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(60, length + 5)

workbook.save(report_path)
print(f"Generated report at {report_path}")
